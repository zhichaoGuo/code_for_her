import openpyxl


def hh_load_excel(path: str, sheet_num: int):
    excel = openpyxl.load_workbook(path)
    print(excel.sheetnames)
    sheet_name = excel.sheetnames[sheet_num]
    table = excel[sheet_name]
    all_data = {}
    year_data = {}
    mouth_data = {}
    for row in range(table.max_row):
        year = str(table.cell(row + 1, 2).value)
        mouth = str(table.cell(row + 1, 3).value)
        day = str(table.cell(row + 1, 4).value)
        value = str(table.cell(row + 1, 5).value)
        # 除倒数第一行
        if row < table.max_row - 1:
            next_year = str(table.cell(row + 2, 2).value)
            next_mouth = str(table.cell(row + 2, 3).value)
            next_day = str(table.cell(row + 2, 4).value)
            next_value = str(table.cell(row + 2, 5).value)
            # 本行和下行同年同月
            if year == next_year:
                if mouth == next_mouth:
                    mouth_data[day] = value
                else:
                    mouth_data[day] = value
                    year_data[mouth] = mouth_data
                    mouth_data = {}
            else:
                all_data[year] = year_data
                year_data = {}
        else:
            mouth_data[day] = value
            year_data[mouth] = mouth_data
            mouth_data = {}
            all_data[year] = year_data
            year_data = {}
    return all_data, sheet_name


def hh_check_data(data):
    pass


class HHData:
    def __init__(self, all_data,sheet_name):
        # 在init中就将类型创建好
        # self.data = {'1999': {'11': {'29': '1.6543',
        #                              '30': '1.6543'
        #                              },
        #                       '12': {'1': '1.6543',
        #                              '2': '1.6543'
        #                              }
        #                       },
        #              '2000': {'1': {'1': '1.6543',
        #                             '2': '1.6543'
        #                             }
        #                       }
        #              }
        self.data = all_data
        self.name = sheet_name

    def __iter__(self):
        return iter(self.data.keys())

    def __getitem__(self, item):
        return HHYearData(self.data[item], item)  # 输入一个年份返回一个年类型

    def __str__(self):
        return self.name

    def year_list(self):
        return list(self.data.keys())

    def sum(self):
        pass


class HHYearData:
    def __init__(self, data: dict, item):
        # self.data = {'11': {'29': '1.6543',
        #                     '30': '1.6543'
        #                     },
        #              '12': {'1': '1.6543',
        #                     '2': '1.6543'
        #                     }
        #              }
        self.data = data
        self.name = item

    def __iter__(self):
        return iter(self.data.keys())

    def __getitem__(self, item):
        return HHMouthData(self.data[item], item)  # 输入一个月份返回一个月类型

    def __str__(self):
        return self.name

    def __eq__(self, other):
        return self.name == other

    def mouth_list(self):
        return list(self.data.keys())

    def sum(self):
        pass

    def avg(self):
        pass

    def max(self):
        pass

    def day_num(self):
        pass


class HHMouthData:
    def __init__(self, data: dict, item):
        # self.data = {'29': '1.6543',
        #              '30': '1.6543'
        #              }
        self.data = data
        self.name = item

    def __iter__(self):
        # 在 for i in HHMouthData: 中使用，i会一次从返回的iter迭代器中依次取值
        return iter(self.data.keys())

    def __getitem__(self, item):
        # 在 HHMouthData[key] 中使用
        return self.data[item]

    def __str__(self):
        # 在直接调用 HHMouthData 时返回 但注意应使用str()强转返回值，不然type为class
        return self.name

    def __eq__(self, other):
        return self.name == other

    def day_list(self):
        return list(self.data.keys())

    def sum(self):
        sum_data = 0
        for value in list(self.data.values()):
            sum_data = sum_data + float(value)
        return sum_data

    def avg(self):
        pass

    def max(self):
        pass

    def day_num(self):
        return len(self.data.keys())


if __name__ == '__main__':
    (all_data,sheet_name) = hh_load_excel(path='source_data_蒸发.xlsx', sheet_num=0)
    # print(sheet_name)
    Data = HHData(all_data,sheet_name)
    for year in Data:
        # 这里筛选年份
        if Data[year] == '1959':
            for mouth in Data[year]:
                # 这里筛选月份
                if Data[year][mouth] == '5':
                    # 这里对数据进行操作
                    print(Data[year][mouth].sum())
