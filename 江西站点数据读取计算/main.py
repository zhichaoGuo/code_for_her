import os
import sys

import yaml
from openpyxl import Workbook

from Utils import *


def return_year_10st_avg_dic(file_name_abs_path):
    """
    通过文件绝对路径名，打开文件，获取第二列年份相同的第十列数值的平均值
    :param file_name_abs_path:文件绝对路径名
    :return:站名，平均值的字典
    eg：站名，{  2011:12
                2012:14}
    """
    day_number = 0
    sum_10st = 0
    avg_10st = {}
    last_year = '0'
    file = open(file_name_abs_path)
    print('打开：' + file_name_abs_path)
    while 1:
        line = file.readline()
        if not line:
            # 计算上一年平均值
            avg_10st.update({last_year: sum_10st / day_number})
            break
        # 将行分割称为列表形如
        # ['58637', '1980', '12', '31', '0.0000', '999999.0000', '999999.0000', '1009.6000', '1.8000', '6.4000', '5.2000', '56.0000', '8.2000', '1005.7000', '-0.2000', '1013.3000', '12.8000', '4.0000', '999003.0000', '35.0000', '\n']
        line_list = line.split(' ')
        # 忽略首行
        if line_list[0] == 'V01301':
            pass
        # 不为首行时执行
        else:
            # txt文件首年执行修正
            if last_year == '0':
                last_year = line_list[1]
            # 在当年内计算
            if line_list[1] == last_year:
                # 判断是否为fake值
                if float(line_list[9]) > 99:
                    pass
                # 不为fake值时，day自加1，sum加上真实值
                else:
                    day_number += 1
                    sum_10st += float(line_list[9])
            # 下一年计算
            else:
                # 计算上一年平均值
                avg_10st.update({last_year: sum_10st / day_number})
                # 更换年
                last_year = line_list[1]
                # 判断是否为fake值
                if float(line_list[9]) > 99:
                    # 如果是fake值不计算当日，将数值初始化
                    day_number = 0
                    sum_10st = 0
                # 不为fake值时，计算当日，并添加数据
                else:
                    day_number = 1
                    sum_10st = float(line_list[9])

    file.close()
    # 取路径名中的文件夹名作为站名
    station_name = os.path.dirname(file_name_abs_path).split('\\')[-1]
    return [station_name, avg_10st]


def gen_excel(result):
    # 新建excel
    wb = Workbook()
    sheet = wb.create_sheet(index=0, title="all station avg")
    print(result)
    for station in result:

        for key in station[1].keys():
            sheet.append([station[0], int(key), station[1][key]])
        # print(station[1])
    wb.save('江西站点年平均温度.xlsx')
    return 1


def gen_new_excel(result):
    '''
    形成  | 站名 | 年份 | 月份 | 日期 | 温度 | 的表格
    :param result:
    :return:
    '''
    wb = Workbook()

    for station in result:
        sheet = wb.create_sheet(title=station[0])
        for i in range(len(station[1])):
            sheet.append([station[0], *station[1][i]])
    del wb['Sheet']
    wb.save('test_remove_same_data.xlsx')
    return 1


def gen_avg_month(result):
    # 需要每个yue平均值
    # 创建文件
    wb = Workbook()
    # 循环每一个站点
    for station in result:
        # 为每个站点建一个sheet
        sheet = wb.create_sheet(title=station[0])
        fist_lin = 1
        month_sum = 0
        month_day = 0
        # 循环每个站点每一天的数据  station[0] 为站名，station[1][i][0] 为年, [1]月，[2]日，[3]为温度
        for i in range(len(station[1])):
            # 判断是否为首行
            if fist_lin == 1:
                fist_lin = 0
                month_sum += station[1][i][3]
                month_day += 1
            # 不是首行
            else:
                # 判断是否为未行
                if i == len(station[1]) - 1:
                    month_sum += station[1][i][3]
                    month_day += 1
                    sheet.append([station[1][i][0], station[1][i][1], month_sum / month_day])
                    fist_lin = 1
                    month_sum = 0
                    month_day = 0
                # 不是最后一行
                else:
                    # 判断当前行是否与上一行日期相同
                    if station[1][i][2] == station[1][i-1][2]:
                        pass
                    else:
                        # 判断当前行是否比上一行日期大，如果大则累加计算月的合
                        if station[1][i][2] > station[1][i - 1][2]:
                            month_sum += station[1][i][3]
                            month_day += 1
                        # 如果不大，则登记已有的平均值,并更新本月值
                        else:
                            sheet.append([station[1][i-1][0], station[1][i-1][1], month_sum / month_day] )
                            month_sum = station[1][i][3]
                            month_day =1


    del wb['Sheet']
    wb.save('avg_month_re.xlsx')
    return 1

def gen_avg_season(result):
    # 创建文件
    wb = Workbook()
    # 循环每一个站点
    for station in result:
        # 为每个站点建一个sheet
        sheet = wb.create_sheet(title=station[0])
        season_sum=[0,0,0,0,0]
        season_day=[0,0,0,0,0]
        # this_winter,this_spring,this_summer,this_autumn,next_winter
        fist_lin = 1
        # 循环每个站点每一天的数据  station[0] 为站名，station[1][i][0] 为年, [1]月，[2]日，[3]为温度
        for i in range(len(station[1])):
            # 判断是否为首行
            if fist_lin == 1:
                fist_lin = 0
                season_sum[return_season(station[1][i][1])] = station[1][i][3]
                season_day[return_season(station[1][i][1])] = 1
            # 不是首行
            else:
                # 判断是否为末行，如果为末行，则添加数据并输出今年的冬春夏秋冬
                if i == len(station[1]) - 1:
                    season_sum[return_season(station[1][i][1])] += station[1][i][3]
                    season_day[return_season(station[1][i][1])] += 1
                    sheet.append([station[1][i - 1][0], '冬季', season_sum[0] / season_day[0]])
                    sheet.append([station[1][i - 1][0], '春季', season_sum[1] / season_day[1]])
                    sheet.append([station[1][i - 1][0], '夏季', season_sum[2] / season_day[2]])
                    sheet.append([station[1][i - 1][0], '秋季', season_sum[3] / season_day[3]])
                    sheet.append([station[1][i - 1][0], '剩余', season_sum[4] / season_day[4]])
                else:
                    # 与上一行是否同一年，如果是，则继续添加数据，如果否则输出上一年的冬春夏秋，并初始化今年的值
                    if station[1][i][0] == station[1][i-1][0]:
                        season_sum[return_season(station[1][i][1])] += station[1][i][3]
                        season_day[return_season(station[1][i][1])] += 1
                    else:
                        try:
                            sheet.append([station[1][i - 1][0], '冬季', season_sum[0] / season_day[0]])
                        except ZeroDivisionError:
                            sheet.append([station[1][i - 1][0], '冬季', 'NODATA'])
                        try:
                            sheet.append([station[1][i - 1][0], '春季', season_sum[1] / season_day[1]])
                        except ZeroDivisionError:
                            sheet.append([station[1][i - 1][0], '春季', 'NODATA'])
                        try:
                            sheet.append([station[1][i - 1][0], '夏季', season_sum[2] / season_day[2]])
                        except ZeroDivisionError:
                            sheet.append([station[1][i - 1][0], '夏季', 'NODATA'])
                        try:
                            sheet.append([station[1][i - 1][0], '秋季', season_sum[3] / season_day[3]])
                        except ZeroDivisionError:
                            sheet.append([station[1][i - 1][0], '秋季', 'NODATA'])
                        tmp = [season_sum[4],season_day[4]]
                        season_sum = [tmp[0], 0, 0, 0, 0]

                        season_day = [tmp[1], 0, 0, 0, 0]
                        season_sum[return_season(station[1][i][1])] += station[1][i][3]
                        season_day[return_season(station[1][i][1])] += 1
    del wb['Sheet']
    wb.save('avg_season.xlsx')
    return 1

if __name__ == '__main__':
    result = []
    file_abs_path_list = get_file_path_list(sys.path[1] + '\\江西')
    # print(file_abs_path_list)
    for file_abs_path in file_abs_path_list:
        txt_name_abs_path_list = get_txt_list(file_abs_path)
        for txt_name_abs_path in txt_name_abs_path_list:
            new_txt_data = return_format_data(txt_name_abs_path)
            if not result:
                result.append(new_txt_data)
            else:
                if new_txt_data[0] == result[-1][0]:
                    for i in range(len(new_txt_data[1])):
                        result[-1][1].append(new_txt_data[1][i])
                else:
                    result.append(new_txt_data)
    # 生成全部数据的excel
    # gen_new_excel(result)
    # 生成月平均数据
    # gen_avg_month(result)
    # 生成季节平均数据
    gen_avg_season(result)
