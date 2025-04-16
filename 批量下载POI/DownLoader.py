import time

import openpyxl
import requests
from pyproj import Transformer

from 批量下载POI.Utils import my_logger


class BaseDownLoader:
    @staticmethod
    def update_status(message):
        print(f"\r{message}", end='', flush=True)

    @staticmethod
    def gcj02_to_wgs84(gcj_lon, gcj_lat):
        # 创建从GCJ-02到WGS-84的转换器
        transformer = Transformer.from_crs("EPSG:4546", "EPSG:4326", always_xy=True)
        wgs_lon, wgs_lat = transformer.transform(gcj_lon, gcj_lat)
        return wgs_lon, wgs_lat

class AMapDownLoader(BaseDownLoader):
    pass

class BaiduDownLoader(BaseDownLoader):
    class Region:
        HeFei="127"
        FuYang="128"
        WuHu="129"
        AnQing="130"
    class Param:
        KeyWord = "query"
        Tag = "tag"
        Region = "region"
        CityLimit = "city_limit"
        PageSize = "page_size"
        PageNum = "page_num"
        Output = "output"
        ApiKey = "ak"
    class KeyWord:
        # https://lbs.baidu.com/index.php?title=open/poitags
        Food = "美食"
        Hotel = "酒店"
        Shopping = "购物"
        Service = "生活服务"
        Company = "公司企业"


    def __init__(self,ak:str):
        self.ak = ak
        self.url = "https://api.map.baidu.com/place/v2/search"
        self.params = {
            BaiduDownLoader.Param.KeyWord: "",
            BaiduDownLoader.Param.Tag: "",
            BaiduDownLoader.Param.Region: BaiduDownLoader.Region.HeFei,
            BaiduDownLoader.Param.CityLimit: True,
            BaiduDownLoader.Param.PageSize: 20,
            BaiduDownLoader.Param.PageNum: 0,
            BaiduDownLoader.Param.Output: "json",
            BaiduDownLoader.Param.ApiKey: self.ak
        }
        self.all_poi_info=[]
        self.start_page = 0
        self.end_page = 1
    def set_params(self,key:str, value):
        self.params[key] = value

    def download(self,key_word:str):
        self.params[BaiduDownLoader.Param.KeyWord] = key_word
        self.start_page = self.params[BaiduDownLoader.Param.PageNum]
        page = self.start_page
        while True:
            self.set_params(BaiduDownLoader.Param.PageNum, page)
            response = requests.get(self.url, params=self.params)
            data = response.json()
            my_logger.debug(f"请求参数：{self.params}")
            my_logger.debug(f"请求结果：{data}")
            # 检查响应状态码
            if response.status_code != 200:
                my_logger.error(f"请求失败，状态码：{response.status_code}")
                break
            # 检查是否有数据返回
            if not data:
                my_logger.error("没有更多数据了")
                break
            # 检查返回的查询状态码
            if data.get("status") != 0:
                my_logger.error(f"查询失败，状态码：{data.get('status')} 原因：{data.get('message')}")
                break
            pois = data.get('results', [])
            if not pois:
                my_logger.error("results 为空")
                break
            self.all_poi_info.extend(pois)
            page += 1
            time.sleep(1)
            if page == 2:
                break
        self.end_page=page
        my_logger.info(f"本次查询开始于第{self.start_page}页，结束于第{self.end_page}页")
        my_logger.info(f"总共获取到{len(self.all_poi_info)}条数据")

    def save_to_csv(self, file_name):
        # 用openpyxl将数据保存到CSV文件
        wb = openpyxl.Workbook()
        sheet_index = 0
        wb.create_sheet(index=sheet_index, title='poi')
        cur_sheet = wb.worksheets[sheet_index]
        cur_row_index = 1
        # 填数据
        for poi in self.all_poi_info:
            cur_col_index = 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("province"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("city"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("area"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("adcode"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("uid"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, poi.get("name"))
            cur_col_index += 1
            location = poi.get("location")
            cur_sheet.cell(cur_row_index, cur_col_index, location.get("lng"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, location.get("lat"))
            cur_col_index += 1
            cur_sheet.cell(cur_row_index, cur_col_index, location.get("address"))
            cur_row_index += 1
        # 保存工作表
        wb.save(f"{self.start_page}_{self.end_page}_{file_name}.csv")
