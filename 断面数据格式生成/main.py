import re
from typing import Optional

import openpyxl
from openpyxl.reader.excel import load_workbook


class SectionDataManager:
    def __init__(self, top_id: str, river_name: str, drag_coefficient: str, pile_fix: str):
        self.top_id = top_id
        self.river_name = river_name
        self.drag_coefficient = drag_coefficient
        self.pile_fix = pile_fix

    @staticmethod
    def _get_mileage(pile_fix: str, pile_number) -> Optional[float]:
        """
        根据桩号返回里程值
        pile_number: str eg:"CH0+018.0"
        return: 18
        """
        if pile_fix not in pile_number:
            print('桩号：%s 不存在前缀：%s' % (pile_number, pile_fix))
            return None
        pattern = rf"{pile_fix}\d+\+\d+\.\d"
        # 使用re.search()函数在字符串中查找第一个匹配的子串
        mileage = re.search(pattern, pile_number)
        # 如果找到了匹配的子串，就将其转换为浮点数并返回
        if mileage:
            mileage_list = mileage.group()[len(pile_fix):].split("+")
            return float(mileage_list[0]) * 1000 + float(mileage_list[1])
            # return float(mileage.group())
        # 如果没有找到匹配的子串，就返回None
        else:
            return None

    @staticmethod
    def _mark_section_data(raw_data: dict, extended_data: list = None):
        """
        标记左岸地平河底高程右岸地平，有拓展数据时使用拓展数据，没有拓展数据时使用排序算法
        """
        mark_default = "<#0>"
        mark_min_h = "<#2>"
        mark_l_max = "<#1>"
        mark_r_max = "<#4>"
        marked_data = []
        mark_list = [mark_l_max, mark_min_h, mark_r_max]
        extended_index = 0
        if extended_data:
            for pre_data_key in raw_data.keys():
                mark = mark_default
                if extended_index < 3:
                    if raw_data[pre_data_key] == extended_data[extended_index]:
                        mark = mark_list[extended_index]
                        extended_index += 1
                marked_data.append([pre_data_key, raw_data[pre_data_key], mark])
        # 在没有辅助数据的情况下使用算法排序
        else:
            min_h = None
            l_max_h = None
            r_max_h = None
            min_h_index = 0
            l_max_h_index = 0
            r_max_h_index = 0
            cur_index = 0
            for pre_data_key in raw_data.keys():
                cur_value = raw_data[pre_data_key]
                if not min_h:
                    min_h = cur_value
                    l_max_h = cur_value
                    r_max_h = cur_value
                else:
                    if cur_value < min_h:
                        if r_max_h > l_max_h:
                            l_max_h = r_max_h
                            l_max_h_index = r_max_h_index
                        r_max_h = cur_value
                        r_max_h_index = cur_index
                        min_h = cur_value
                        min_h_index = cur_index
                    else:
                        if r_max_h < cur_value:
                            r_max_h = cur_value
                            r_max_h_index = cur_index
                marked_data.append([pre_data_key, cur_value, mark_default])
                cur_index += 1
            marked_data[l_max_h_index][2] = mark_l_max
            marked_data[min_h_index][2] = mark_min_h
            marked_data[r_max_h_index][2] = mark_r_max

        return marked_data

    def gen_section_data(self, pile_number: str, raw_data: dict, extended_data: dict = {}) -> list:
        """
        根据原始数据，生成断面数据
        pile_number: str eg:"CH0+018.0"
        raw_data: dict eg:{-80.5:44.05,-75.0:45.3}
        return: [
                    [],
                    [],
                    []
                    ...
                ]
        """
        mileage = SectionDataManager._get_mileage(self.pile_fix, pile_number)
        if not mileage:
            print('没有依照桩号规则匹配到内容，请中断程序，并检查桩号：%s ' % pile_number)
            return []
        data = [
            [self.top_id],
            [self.river_name],
            [mileage],
            ["COORDINATES"],
            ["0"],
            ["FLOW DIRECTION"],
            ["0"],
            ["DATUM"],
            ["0"],
            ["RADIUS TYPE"],
            ["0"],
            ["DIVIDE X-Section"],
            ["0"],
            ["SECTION ID"],
            [pile_number],
            ["INTERPOLATED"],
            ["0"],
            ["ANGLE"],
            ["0.00   0"],
            ["PROFILE  ", len(raw_data)]
        ]
        marked_data = SectionDataManager._mark_section_data(raw_data, extended_data.get(mileage))
        data.extend(marked_data)
        data.append(["*****************************"])
        return data


def load_section_data_excel(excel_path: str) -> dict:
    """
    读取断面原始数据excel文件
    | 断面号1 |     |
    | x      | H   |
    | ...    | ... |
    | 断面号2 |     |
    | ...    |     |
    data = {
        "CH0+256.0":
                    {   -80.5:44.05,
                        -75.0:45.3,
                        ...,
                        30:48.45
                    },
        "CH0+756.0":
                    {
                        ...
                    }
            }

    """
    data = {}
    wb = load_workbook(excel_path)
    ws = wb.worksheets[0]
    mileage = ws.cell(1, 1).value
    for i in range(ws.max_row):
        first = ws.cell(row=i + 1, column=1).value
        second = ws.cell(i + 1, 2).value
        if second is None:
            if first is not None:
                mileage = first
                data[mileage] = {}
        else:
            data[mileage][first] = second
    return data


def load_extended_data_excel(excel_path: str) -> dict:
    """
    读取拓展胡数据excel文件
    | 里程 | 左地平 | 河底高程 | 右地平 |
    | --- | --- | --- | --- |
    data = {
            "0":[Null,0,Null],
            "256.0":[46,32.5,60]
            }
    """
    data = {}
    wb = load_workbook(excel_path)
    ws = wb.worksheets[0]
    for i in range(ws.max_row):
        if i == 0:
            continue
        mileage = ws.cell(i + 1, 1).value
        l_max = ws.cell(i + 1, 2).value
        min_h = ws.cell(i + 1, 3).value
        r_max = ws.cell(i + 1, 4).value
        data[mileage] = [l_max, min_h, r_max]
    return data


def save_file(file_path: str, file_data: list):
    wb = openpyxl.Workbook()
    sheet_index = 0
    wb.create_sheet(index=sheet_index, title='xyz')
    cur_sheet = wb.worksheets[sheet_index]
    for row_index in range(len(file_data)):
        for col_index in range(len(file_data[row_index])):
            cur_sheet.cell(row_index + 1, col_index + 1, file_data[row_index][col_index])
    wb.save(file_path)


def main():
    file_path = "./断面原始数据.xlsx"  # 原始断面数据excel
    extended_data_path = "./拓展数据.xlsx"  # 拓展断面数据excel
    top_id = "Topo001"  # top id
    river_name = "Branch1"  # 河流名称
    drag_coefficient = "0.03"  # 阻力系数
    pile_fix = "CH"  # 桩号前缀
    # 读取原始断面数据
    section_data = load_section_data_excel(excel_path=file_path)
    # 读取拓展数据
    extended_data = load_extended_data_excel(excel_path=extended_data_path)
    # print(section_data)
    # print(extended_data)
    # 断面数据管理器
    section_manager = SectionDataManager(top_id, river_name, drag_coefficient, pile_fix)
    # #
    xyz_data = []
    for pre_mileage in section_data.keys():
        # xyz_data.extend(section_manager.gen_section_data(pre_mileage, section_data[pre_mileage], extended_data))
        xyz_data.extend(section_manager.gen_section_data(pre_mileage, section_data[pre_mileage]))
    save_file("./test.xlsx", xyz_data)
    print(xyz_data)


if __name__ == '__main__':
    main()
